#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fetch_kosei.py — 関東信越厚生局「届出受理医療機関名簿」から
泌尿器科関連の施設基準を抽出し、紹介先ファインダー用 facilities.json を生成する。

使い方:
  pip install openpyxl requests
  python3 fetch_kosei.py                # 取得→抽出→facilities.json
  python3 fetch_kosei.py --local 名簿.xlsx   # 手動DLしたExcelを使う
  python3 fetch_kosei.py --no-geocode   # ジオコーディングを省略（区中心座標）

出力: facilities.json（アプリの FACILITIES と同スキーマ, dir="up", real=true）

注意:
- 厚生局のページ構成・Excelレイアウトは変わることがある。初回はログを見て確認。
- 厚生局サイトは海外IPを遮断している模様。海外環境（一部CI含む）では
  ダウンロードに失敗する → 手動DLして --local を使う。
"""
import argparse, json, re, sys, time, unicodedata
from pathlib import Path

try:
    import requests
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl requests を先に実行してください")

# ============================== 設定 ==============================

# 2026-07時点: 届出受理医療機関名簿は kijyun.html に掲載。
# Excelは都県別ではなく「医科（ZIP）」1本（shisetsu_ika_*.zip、中に都県別xlsx）。
LIST_PAGE = "https://kouseikyoku.mhlw.go.jp/kantoshinetsu/chousa/kijyun.html"
ZIP_LINK_RE = r'href="([^"]*shisetsu_ika[^"]*\.zip)"'   # 医科の名簿ZIP
ZIP_MEMBER_PREF = "13"   # ZIP内の東京都ファイル（都道府県コード接頭辞）

# 対象エリア（所在地の部分一致）。空リストなら都内全域。
AREA_FILTER = ["葛飾区", "足立区", "江戸川区", "墨田区", "荒川区"]

# 施設基準キーワード → アプリのcap
# all: すべて含む / any: どれか含む（届出名称に対して部分一致・正規化後）
# 監修メモ（2026-07 泌尿器科医確認済み）:
# - holep / tul / tot は施設基準届出が不要な手技のため名簿に載らない
#   → 自動判定を諦め、医師が手動入力する別枠capとして扱う（ここには書かない）
CAP_RULES = [
    {"cap": "robot",   # ロボット支援手術（前立腺・腎・尿管・腎盂・膀胱）
     "all": ["内視鏡手術用支援機器"],
     "any": ["前立腺悪性腫瘍", "腎悪性腫瘍", "膀胱悪性腫瘍", "腎盂", "尿管"]},
    {"cap": "eswl",    # 体外衝撃波腎・尿管結石破砕術
     "any": ["体外衝撃波腎・尿管結石破砕術"]},
    {"cap": "aus",     # 人工尿道括約筋植込・置換術（男性高度尿失禁）
     "any": ["人工尿道括約筋"]},
    {"cap": "mrifusion",  # MRI・超音波融合画像下前立腺針生検
     "any": ["前立腺針生検"]},
    {"cap": "snm",     # 仙骨神経刺激（過活動膀胱）※便失禁のみの届出は除外
     "all": ["仙骨神経刺激"],
     "any": ["過活動膀胱"]},
    {"cap": "microwave",  # 経皮的前立腺がんマイクロ波焼灼・凝固療法
     "any": ["経皮的前立腺がんマイクロ波焼灼"]},
    {"cap": "rfa",     # 腎悪性腫瘍ラジオ波焼灼療法（副腎腫瘍RFAは対象外）
     "any": ["腎悪性腫瘍ラジオ波焼灼"]},
    {"cap": "hydro",   # 膀胱水圧拡張術（ハンナ型間質性膀胱炎）
     "any": ["膀胱水圧拡張"]},
    {"cap": "tese",    # 精巣内精子採取術
     "any": ["精巣内精子採取"]},
    {"cap": "lsc",     # 腹腔鏡下仙骨腟固定術（ロボット支援含む）
     "any": ["腹腔鏡下仙骨腟固定術"]},
    {"cap": "rasc",    # ロボット支援仙骨腟固定術
     "all": ["仙骨腟固定術", "内視鏡手術用支援機器"]},
]

# 手動監修情報（施設名の部分一致でマージ）。毎月の自動再生成でも消えない。
# caps: 公式サイトの診療内容・手術実績で確認した届出外の対応手技（2026-07確認）
# ★ここが臨床監修の本体その2。年1回程度は各院HPと突き合わせて見直すこと。
MANUAL_INFO = {
    "イムス東京葛飾総合病院": {
        "caps": ["holep", "tul", "tot", "rezum", "pul", "pnl", "rasc"],
        "url": "https://ims-tokyo-katsushika.com/"},
    "葛飾医療センター": {"caps": ["holep", "tueb", "tul", "botox", "pnl"],
        "url": "https://www.jikei.ac.jp/hospital/katsushika/"},
    "東部地域病院": {"caps": ["holep", "turp", "tul", "rezum"],
        "url": "https://www.tmhp.jp/tobu/"},
    "同愛記念病院": {"caps": ["holep", "tueb", "tul", "rezum"],
        "url": "https://www.doai.jp/"},
    "足立医療センター": {"caps": ["tul", "turp", "pul"],
        "url": "https://twmu-amc.jp/"},
    "墨東病院": {"caps": ["tul"], "url": "https://www.tmhp.jp/bokutoh/"},
    "平成立石病院": {"caps": ["tul"], "url": "https://www.heisei-tateishi.net/"},
    "東京臨海病院": {"caps": ["tul"], "url": "https://www.tokyorinkai.jp/"},
    "博慈会記念総合病院": {"caps": ["tul", "tot"],
        "url": "https://hakujikai.or.jp/kinen/"},
    "森山記念病院": {"caps": ["tul"], "url": "https://mk.moriyamaikai.or.jp/"},
    "はせがわ病院": {"caps": [], "url": "https://seishukai.or.jp/"},
    "江戸川病院": {"caps": [], "url": "https://www.edogawa.or.jp/"},
}

# 下り（かかりつけ・逆紹介先）。基本は down_facilities.json（fetch_navii が生成、ナビイ由来の
# 全院＋uro_level/caps/座標）を取り込み、以下の手動データをマージして facilities.json に連結する。
DOWN_JSON = Path("down_facilities.json")

# ① Navii に載らない実在の下り施設（手動フル管理）。病院や未登録院はここで補う。
#   ★臨床監修の本体その3。年1回はHPと突き合わせて見直すこと。
HAND_DOWN_FULL = [
    {"id": "dn001", "name": "立石駅前泌尿器・内科外科クリニック", "dir": "down",
     "real": True, "lat": 35.736588, "lng": 139.845261,
     "addr": "葛飾区立石1-7-25 1F", "tel": "03-5670-3366",
     "url": "https://tateishi-urology.com/",
     "station": "京成立石駅", "walk": 4, "bus": "", "bf": True,
     "hours": "土曜午後も診療（HP参照）",
     "caps": ["keizoku", "rezum"], "uro_level": "専門",
     "source": "公式サイト（日帰りWAVE治療・泌尿器一般）"},
    {"id": "dn003", "name": "井口腎泌尿器科・内科 新小岩", "dir": "down",
     "real": True, "lat": 35.716011, "lng": 139.860214,
     "addr": "葛飾区新小岩1-49-10", "tel": "03-6231-5931",
     "url": "https://www.iguchi-jinhinyoki-shinkoiwa.jp/",
     "station": "新小岩駅", "walk": None, "bus": "", "bf": True,
     "hours": "HP参照（人工透析併設）",
     "caps": ["keizoku"], "uro_level": "専門",
     "source": "公式サイト（泌尿器一般・透析）"},
    {"id": "dn005", "name": "金町中央病院", "dir": "down",
     "real": True, "lat": 35.763023, "lng": 139.865677,
     "addr": "葛飾区金町1-9-1", "tel": "03-3607-2001",
     "url": "https://www.reiroukai.or.jp/",
     "station": "金町駅", "walk": None, "bus": "", "bf": True,
     "hours": "HP参照",
     "caps": ["keizoku"], "uro_level": "専門",
     "source": "公式サイト（泌尿器科外来：排尿障害・結石・前立腺がん等）"},
    # ナビイ未収載（2023-6開院・報告なし）。医師の知見で追加。座標は同ビル(三鈴ビル)登録院を流用。
    {"id": "dn007", "name": "北千住セブンデイズクリニック", "dir": "down",
     "real": True, "lat": 35.750776, "lng": 139.802423,
     "addr": "足立区千住3-33 三鈴ビル1F", "tel": "要確認",
     "url": "https://7daysclinic.jp/",
     "station": "北千住駅", "walk": 2, "bus": "", "bf": True,
     "hours": "週7日診療（HP参照）",
     "caps": ["keizoku"], "uro_level": "対応",   # 専門度/capは暫定。院長専門・膀胱鏡等が分かれば更新
     "source": "公式サイト（内科・皮膚科・泌尿器科・性病科／週7日診療）"},
]

# ② Navii由来の院に上書きする手動情報（名称の全キーワード一致でマージ）。
#   駅・徒歩・電話など、Naviiに無い/精度が低い項目を補う。caps/uro_level は上書きしない。
HAND_OVERRIDES = [
    {"match": ["ばんどう"], "station": "堀切菖蒲園駅", "walk": 1,
     "tel": "03-6662-8255", "hours": "HP参照"},
    {"match": ["井口", "亀有"], "station": "亀有駅", "walk": 1,
     "tel": "03-3838-8721", "hours": "HP参照（人工透析併設）"},
]

def load_downs():
    """down_facilities.json（ナビイ由来）に手動オーバーレイを当て、手動フル院を足して返す。"""
    downs = []
    if DOWN_JSON.exists():
        downs = json.loads(DOWN_JSON.read_text(encoding="utf-8"))
        for nd in downs:
            for ov in HAND_OVERRIDES:
                if all(k in nd.get("name", "") for k in ov["match"]):
                    for f in ("station", "walk", "hours", "tel", "url"):
                        if f in ov and ov[f] not in (None, ""):
                            nd[f] = ov[f]
        print(f"  下り: ナビイ {len(downs)} 件 ＋ 手動 {len(HAND_DOWN_FULL)} 件")
    else:
        print("  [warn] down_facilities.json が無い → 手動下りのみ（fetch_navii を先に実行）")
    return downs + HAND_DOWN_FULL

GSI_GEOCODER = "https://msearch.gsi.go.jp/address-search/AddressSearch?q="
GEOCODE_CACHE = Path("geocode_cache.json")
OUTPUT = Path("facilities.json")

# ジオコーディング失敗時のフォールバック（区の概略中心）
WARD_CENTERS = {
    "葛飾区": (35.7434, 139.8472), "足立区": (35.7750, 139.8046),
    "江戸川区": (35.7068, 139.8683), "墨田区": (35.7107, 139.8016),
    "荒川区": (35.7362, 139.7830), "東京都": (35.6895, 139.6917),
}

UA = {"User-Agent": "referral-finder-data/1.0 (clinical tool; contact site owner)"}

# ============================== 取得 ==============================

def norm(s):
    """全角半角ゆらぎを吸収して比較用に正規化"""
    return unicodedata.normalize("NFKC", str(s or "")).replace(" ", "").replace("　", "")

def download_excels():
    """kijyun.html から医科名簿ZIPを取得し、東京都のxlsxを展開して返す。"""
    import io, zipfile
    print(f"[1/4] 一覧ページ取得: {LIST_PAGE}")
    r = requests.get(LIST_PAGE, headers=UA, timeout=30)
    r.raise_for_status()
    m = re.search(ZIP_LINK_RE, r.text)
    if not m:
        sys.exit("医科名簿ZIPリンクが見つからない。ページ構成が変わった可能性 → LIST_PAGE を開いて確認し、--local で手動指定を。")
    url = requests.compat.urljoin(LIST_PAGE, m.group(1))
    print(f"  ZIP DL: {url}")
    d = requests.get(url, headers=UA, timeout=300)
    d.raise_for_status()
    zf = zipfile.ZipFile(io.BytesIO(d.content))
    members = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
    tokyo = [n for n in members
             if Path(n).name.startswith(ZIP_MEMBER_PREF) or "東京" in n]
    if not tokyo:
        sys.exit(f"ZIP内に東京都のxlsxが見つからない。members={members}")
    paths = []
    for i, n in enumerate(tokyo):
        p = Path(f"meibo_{i}.xlsx")
        p.write_bytes(zf.read(n))
        print(f"  展開 → {p.name}  ({n})")
        paths.append(p)
    return paths

# ============================== パース ==============================

# ヒントは優先順（先頭ほど優先）。「所在地（郵便番号）」列があるため
# addr は「住所」を先に探す。
HEADER_HINTS = {
    "name":  ["医療機関名称", "医療機関名", "名称"],
    "addr":  ["住所", "所在地"],
    "tel":   ["電話"],
    "todoke":["受理届出名称", "届出名称", "届出"],
}

def detect_columns(ws):
    """先頭30行からヘッダ行と列位置を推定（ヒント優先順で列を選ぶ）"""
    for r in range(1, 31):
        row = [norm(c.value) for c in ws[r]]
        cols = {}
        for key, hints in HEADER_HINTS.items():
            for h in hints:                      # 優先順に走査
                hit = next((i for i, v in enumerate(row) if v and h in v), None)
                if hit is not None:
                    cols[key] = hit
                    break
        if {"name", "addr", "todoke"} <= cols.keys():
            return r, cols
    return None, None

def parse_workbook(path):
    """1ブック分の (name, addr, tel, 届出名称) レコードを返す"""
    print(f"[2/4] 解析: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []
    for ws in wb.worksheets:
        hdr, cols = detect_columns(ws)
        if not cols:
            continue
        print(f"  シート'{ws.title}' ヘッダ{hdr}行目 列={cols}")
        last = {"name": "", "addr": "", "tel": ""}
        for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
            def cell(k):
                i = cols.get(k)
                return norm(row[i]) if i is not None and i < len(row) else ""
            name, addr, tel, tod = cell("name"), cell("addr"), cell("tel"), cell("todoke")
            # 名簿は同一機関の2行目以降が空欄のことがある → 直前値を引き継ぐ
            if name: last = {"name": name, "addr": addr or last["addr"],
                             "tel": tel or last["tel"]}
            elif not tod:
                continue
            rec = {**last, "todoke": tod}
            if addr: rec["addr"] = addr
            if tel:  rec["tel"] = tel
            if rec["name"] and rec["todoke"]:
                records.append(rec)
    print(f"  届出行 {len(records)} 件")
    return records

# ============================== 抽出・変換 ==============================

def match_caps(todoke):
    caps = set()
    t = norm(todoke)
    for rule in CAP_RULES:
        ok_all = all(k in t for k in rule.get("all", []))
        anys = rule.get("any", [])
        ok_any = (not anys) or any(k in t for k in anys)
        if ok_all and ok_any:
            caps.add(rule["cap"])
    return caps

def ward_of(addr):
    for w in WARD_CENTERS:
        if w in addr:
            return w
    return "東京都"

def geocode(addr, cache, enabled=True):
    if addr in cache:
        return cache[addr]
    latlng = None
    # 名簿の住所は「千代田区…」のように都名を欠くため補う
    q = addr if addr.startswith("東京都") else "東京都" + addr
    if enabled:
        try:
            r = requests.get(GSI_GEOCODER + requests.utils.quote(q),
                             headers=UA, timeout=15)
            j = r.json()
            if j:
                lng, lat = j[0]["geometry"]["coordinates"]
                latlng = [lat, lng]
        except Exception:
            latlng = None
        time.sleep(0.25)  # GSIに配慮
    if latlng is None:
        lat, lng = WARD_CENTERS[ward_of(addr)]
        latlng = [lat, lng, "approx"]
    cache[addr] = latlng
    return latlng

def build_facilities(records, do_geocode=True):
    print("[3/4] 施設集約・cap変換")
    fac = {}
    for r in records:
        if AREA_FILTER and not any(a in r["addr"] for a in AREA_FILTER):
            continue
        caps = match_caps(r["todoke"])
        if not caps:
            continue
        key = (r["name"], r["addr"])
        f = fac.setdefault(key, {"caps": set(), "tel": r["tel"], "todoke": []})
        f["caps"] |= caps
        f["todoke"].append(r["todoke"])

    cache = json.loads(GEOCODE_CACHE.read_text()) if GEOCODE_CACHE.exists() else {}
    out = []
    for i, ((name, addr), f) in enumerate(sorted(fac.items())):
        ll = geocode(addr, cache, do_geocode)
        approx = len(ll) == 3
        # 手動監修情報のマージ（部分一致）
        manual_caps, url = [], ""
        for key, info in MANUAL_INFO.items():
            if key in name:
                manual_caps = info.get("caps", [])
                url = info.get("url", "")
                break
        entry = {
            "id": f"up{i:03d}",
            "name": name,
            "dir": "up",
            "real": True,
            "lat": ll[0], "lng": ll[1],
            "addr": addr,
            "tel": f["tel"] or "要確認",
            "url": url,
            "station": "", "walk": None,     # 駅情報は次フェーズ
            "bus": "", "bf": True,
            "hours": "紹介予約制（要確認）" + ("／位置は概算" if approx else ""),
            "caps": sorted(f["caps"] | set(manual_caps)),
            "source": "関東信越厚生局 届出受理医療機関名簿",
            "todoke": sorted(set(f["todoke"])),  # 監査用に元の届出名称を残す
        }
        if manual_caps:
            entry["manual_caps"] = sorted(set(manual_caps))
            entry["source"] += "＋公式サイト（手動監修）"
        out.append(entry)
    GEOCODE_CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=1))
    downs = load_downs()
    out += downs
    print(f"  施設 {len(out)} 件（上り{len(out)-len(downs)}＋下り{len(downs)}）")
    return out

# ============================== main ==============================

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--local", nargs="*", help="手動DLしたExcelパス")
    ap.add_argument("--no-geocode", action="store_true")
    ap.add_argument("--dump-names", action="store_true",
                    help="届出名称の一覧だけ出力（CAP_RULES監修用）")
    a = ap.parse_args()

    paths = [Path(p) for p in a.local] if a.local else download_excels()
    records = []
    for p in paths:
        records += parse_workbook(p)

    if a.dump_names:
        names = sorted({r["todoke"] for r in records})
        Path("todoke_names.txt").write_text("\n".join(names), encoding="utf-8")
        print(f"届出名称 {len(names)} 種 → todoke_names.txt（CAP_RULESの監修に）")
        return

    out = build_facilities(records, do_geocode=not a.no_geocode)
    OUTPUT.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"[4/4] 出力 → {OUTPUT}（アプリの DATA_URL に置く）")

if __name__ == "__main__":
    main()
